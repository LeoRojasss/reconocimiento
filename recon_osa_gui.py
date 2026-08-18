"""
Interfaz grafica para el cruce CONSULTA OSA vs catalogo Edimusica.

Flujo:
  1. Elegir el archivo de la OSA y el archivo de obras de Edimusica.
  2. Se analizan automaticamente (coincidencias seguras vs dudosas).
  3. Las coincidencias dudosas se revisan una por una en modo "juego"
     (botones grandes / flechas del teclado: si coincide o no coincide).
  4. Se genera el archivo final para enviar a la OSA, modificando solo las
     obras identificadas (automaticas + confirmadas a mano). El archivo
     original de la OSA nunca se modifica.

El progreso de la revision se guarda automaticamente para poder cerrar el
programa y continuar despues sin perder lo ya decidido.
"""

import json
import sys
import threading
import time
import webbrowser
from collections import deque
from pathlib import Path
from tkinter import StringVar, filedialog, messagebox

import customtkinter as ctk
import openpyxl
from PIL import Image

import matching_core as mc


def app_dir():
    """Carpeta del script, o del .exe cuando esta empaquetado con PyInstaller."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


SESSIONS_DIR = app_dir() / "sesiones_revision"
SESSIONS_DIR.mkdir(exist_ok=True)

ASSETS_DIR = app_dir() / "assets"
OSA_LOGO_PATH = ASSETS_DIR / "osa_logo.png"
FUENTES_LOGO_PATH = ASSETS_DIR / "fuentes_logo.jpg"


def load_logo(path, height):
    """Carga un logo manteniendo su proporcion original, escalado a `height` px."""
    try:
        img = Image.open(path)
    except Exception:
        return None
    w, h = img.size
    return ctk.CTkImage(light_image=img, dark_image=img, size=(int(height * w / h), height))


def session_path_for(osa_path: Path) -> Path:
    return SESSIONS_DIR / f"{osa_path.stem}.json"


# ---------------------------------------------------------------------------
# Paleta y tipografia
# ---------------------------------------------------------------------------

BG = "#F5F6FA"
CARD = "#FFFFFF"
BORDER = "#E5E7EB"

PRIMARY = "#4F46E5"
PRIMARY_HOVER = "#4338CA"
PRIMARY_LIGHT = "#EEF2FF"

SUCCESS = "#16A34A"
SUCCESS_HOVER = "#15803D"
SUCCESS_LIGHT = "#ECFDF5"

DANGER = "#DC2626"
DANGER_HOVER = "#B91C1C"
DANGER_LIGHT = "#FEF2F2"

WARNING = "#D97706"
WARNING_LIGHT = "#FFFBEB"

MUTED = "#6B7280"
MUTED_LIGHT = "#F3F4F6"
TEXT = "#111827"
TEXT_SECONDARY = "#4B5563"

F = "Segoe UI"
FONT_APP_TITLE = (F, 16, "bold")
FONT_APP_SUB = (F, 11)
FONT_H1 = (F, 20, "bold")
FONT_H2 = (F, 14, "bold")
FONT_BODY = (F, 12)
FONT_BODY_BOLD = (F, 12, "bold")
FONT_SMALL = (F, 10)
FONT_CARD_TITLE = (F, 21, "bold")
FONT_BTN = (F, 13, "bold")
FONT_BTN_BIG = (F, 14, "bold")
FONT_STAT_NUM = (F, 30, "bold")
FONT_MONO = ("Consolas", 10)


def card(parent, **kwargs):
    opts = dict(fg_color=CARD, corner_radius=14, border_width=1, border_color=BORDER)
    opts.update(kwargs)
    return ctk.CTkFrame(parent, **opts)


def badge(parent, text, fg, bg):
    return ctk.CTkLabel(parent, text=text, font=FONT_SMALL, text_color=fg,
                         fg_color=bg, corner_radius=8, padx=10, pady=3)


def h_button(parent, text, command, fg=PRIMARY, hover=PRIMARY_HOVER, text_color="white",
             font=FONT_BTN, height=40, corner_radius=10, border_width=0, border_color=None,
             state="normal"):
    return ctk.CTkButton(
        parent, text=text, command=command, fg_color=fg, hover_color=hover,
        text_color=text_color, font=font, height=height, corner_radius=corner_radius,
        border_width=border_width, border_color=border_color or fg, state=state,
    )


def ghost_button(parent, text, command, text_color=TEXT_SECONDARY, hover=MUTED_LIGHT, font=FONT_BODY_BOLD):
    return ctk.CTkButton(
        parent, text=text, command=command, fg_color="transparent", hover_color=hover,
        text_color=text_color, font=font, height=32, corner_radius=8,
    )


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")

        self.title("Reconocimiento OSA - Edimusica")
        self.geometry("1080x780")
        self.minsize(960, 700)
        self.configure(fg_color=BG)

        self.osa_path = None
        self.db_path = None
        self.match_index = None
        self.wb = None
        self.ws = None
        self.scan_results = None
        self.review_groups = []
        self.decisions = {}
        self.session_file = None
        self.queue = deque()
        self.history = []
        self._dots_job = None
        self._dots_n = 0

        self.logo_osa_header = load_logo(OSA_LOGO_PATH, 26)
        self.logo_fuentes_header = load_logo(FUENTES_LOGO_PATH, 40)
        self.logo_osa_card = load_logo(OSA_LOGO_PATH, 34)
        self.logo_fuentes_card = load_logo(FUENTES_LOGO_PATH, 48)

        self._build_header()

        self.content = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        self.content.pack(fill="both", expand=True, padx=28, pady=(20, 24))

        self.show_inicio()

    # -----------------------------------------------------------------
    # chrome / utilidades comunes
    # -----------------------------------------------------------------

    def _build_header(self):
        header = ctk.CTkFrame(self, fg_color=PRIMARY, corner_radius=0, height=68)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)

        left = ctk.CTkFrame(header, fg_color="transparent")
        left.pack(side="left", padx=20, pady=10)

        if self.logo_fuentes_header:
            logo_box = ctk.CTkFrame(left, fg_color="white", corner_radius=9)
            logo_box.pack(side="left", padx=(0, 12))
            ctk.CTkLabel(logo_box, text="", image=self.logo_fuentes_header).pack(padx=7, pady=7)
        else:
            ctk.CTkLabel(left, text="🎵", font=(F, 22), text_color="white").pack(side="left", padx=(0, 10))

        text_col = ctk.CTkFrame(left, fg_color="transparent")
        text_col.pack(side="left")
        ctk.CTkLabel(text_col, text="Reconocimiento OSA", font=FONT_APP_TITLE,
                     text_color="white", anchor="w").pack(anchor="w")
        ctk.CTkLabel(text_col, text="Catalogo Edimusica", font=FONT_APP_SUB,
                     text_color="#DDD9FF", anchor="w").pack(anchor="w")

        if self.logo_osa_header:
            right = ctk.CTkFrame(header, fg_color="white", corner_radius=9)
            right.pack(side="right", padx=20, pady=14)
            ctk.CTkLabel(right, text="", image=self.logo_osa_header).pack(padx=10, pady=6)

    def clear(self):
        self._stop_dots()
        for w in self.content.winfo_children():
            w.destroy()

    def _start_dots(self, base_text, target_var):
        self._dots_n = 0

        def tick():
            self._dots_n = (self._dots_n + 1) % 4
            target_var.set(base_text + "." * self._dots_n)
            self._dots_job = self.after(400, tick)

        tick()

    def _stop_dots(self):
        if self._dots_job:
            try:
                self.after_cancel(self._dots_job)
            except Exception:
                pass
            self._dots_job = None

    def save_session(self):
        if not self.session_file:
            return
        data = {
            "osa_path": str(self.osa_path),
            "db_path": str(self.db_path),
            "updated": time.strftime("%Y-%m-%d %H:%M:%S"),
            "decisions": self.decisions,
        }
        tmp = self.session_file.with_suffix(".json.tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        tmp.replace(self.session_file)

    # -----------------------------------------------------------------
    # Pantalla 1: inicio / seleccion de archivos
    # -----------------------------------------------------------------

    def show_inicio(self):
        self.clear()

        ctk.CTkLabel(self.content, text="Vamos a identificar tus obras",
                     font=FONT_H1, text_color=TEXT, anchor="w").pack(anchor="w")
        ctk.CTkLabel(
            self.content,
            text="Elige el archivo de la OSA y tu catalogo de obras. El resto es automatico.",
            font=FONT_BODY, text_color=MUTED, anchor="w",
        ).pack(anchor="w", pady=(2, 22))

        self.osa_var = StringVar(value=str(self.osa_path) if self.osa_path else "")
        self.db_var = StringVar(value=str(self.db_path) if self.db_path else "")

        self._file_card("📄", "Archivo CONSULTA OSA", "Formato .xlsx que envia la OSA cada trimestre",
                         self.osa_var, self._pick_osa)
        self._file_card("🗂️", "Archivo Obras Edimusica", "Tu catalogo propio: TITULO, NOMAUTOR, % ...",
                         self.db_var, self._pick_db)

        self.btn_analizar = h_button(self.content, "Analizar archivos  →", self._start_analysis,
                                      height=48, font=FONT_BTN_BIG, state="disabled")
        self.btn_analizar.pack(anchor="w", pady=(20, 8))

        self.status_var = StringVar(value="")
        ctk.CTkLabel(self.content, textvariable=self.status_var, font=FONT_BODY,
                     text_color=MUTED, anchor="w").pack(anchor="w")

    def _file_card(self, icon, title, subtitle, var, command):
        c = card(self.content)
        c.pack(fill="x", pady=8)
        inner = ctk.CTkFrame(c, fg_color="transparent")
        inner.pack(fill="x", padx=18, pady=16)

        top = ctk.CTkFrame(inner, fg_color="transparent")
        top.pack(fill="x")
        ctk.CTkLabel(top, text=icon, font=(F, 20)).pack(side="left", padx=(0, 10))
        col = ctk.CTkFrame(top, fg_color="transparent")
        col.pack(side="left", fill="x", expand=True)
        ctk.CTkLabel(col, text=title, font=FONT_H2, text_color=TEXT, anchor="w").pack(anchor="w")
        ctk.CTkLabel(col, text=subtitle, font=FONT_SMALL, text_color=MUTED, anchor="w").pack(anchor="w")
        h_button(top, "Elegir archivo", command, fg=CARD, hover=PRIMARY_LIGHT, text_color=PRIMARY,
                 font=FONT_BODY_BOLD, height=34, corner_radius=8, border_width=1).pack(side="right")

        path_box = ctk.CTkFrame(inner, fg_color=MUTED_LIGHT, corner_radius=8)
        path_box.pack(fill="x", pady=(12, 0))
        path_lbl = ctk.CTkLabel(path_box, textvariable=var, font=FONT_SMALL, text_color=TEXT_SECONDARY,
                                 anchor="w", justify="left")
        path_lbl.pack(fill="x", padx=10, pady=8)
        if not var.get():
            var.set("Sin seleccionar")

    def _pick_osa(self):
        path = filedialog.askopenfilename(
            title="Selecciona el archivo de la CONSULTA OSA",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=str(app_dir()),
        )
        if path:
            self.osa_path = Path(path)
            self.osa_var.set(str(self.osa_path))
            self._update_analyze_btn_state()

    def _pick_db(self):
        path = filedialog.askopenfilename(
            title="Selecciona el archivo de Obras Edimusica",
            filetypes=[("Excel", "*.xlsx")],
            initialdir=str(app_dir()),
        )
        if path:
            self.db_path = Path(path)
            self.db_var.set(str(self.db_path))
            self._update_analyze_btn_state()

    def _update_analyze_btn_state(self):
        ok = bool(self.osa_path and self.db_path)
        self.btn_analizar.configure(state="normal" if ok else "disabled")

    def _start_analysis(self):
        self.btn_analizar.configure(state="disabled")
        self._start_dots("Analizando", self.status_var)
        threading.Thread(target=self._run_analysis, daemon=True).start()

    def _run_analysis(self):
        try:
            match_index = mc.MatchIndex(self.db_path)
            wb = openpyxl.load_workbook(self.osa_path, data_only=False)
            ws = wb[mc.OSA_SHEET]
            scan_results = mc.scan_osa_rows(ws, match_index)
            review_groups = mc.build_review_groups(scan_results)
        except Exception as exc:  # noqa: BLE001
            self.after(0, self._analysis_failed, str(exc))
            return
        self.after(0, self._analysis_done, match_index, wb, ws, scan_results, review_groups)

    def _analysis_failed(self, message):
        self._stop_dots()
        self.btn_analizar.configure(state="normal")
        self.status_var.set("")
        messagebox.showerror("Error al analizar", message)

    def _analysis_done(self, match_index, wb, ws, scan_results, review_groups):
        self._stop_dots()
        self.match_index = match_index
        self.wb = wb
        self.ws = ws
        self.scan_results = scan_results
        self.review_groups = review_groups
        self.decisions = {}

        self.session_file = session_path_for(self.osa_path)
        if self.session_file.exists():
            try:
                data = json.loads(self.session_file.read_text(encoding="utf-8"))
                n = len(data.get("decisions", {}))
                if n and messagebox.askyesno(
                    "Sesion encontrada",
                    f"Ya hay {n} decisiones guardadas para este archivo "
                    f"(ultima actualizacion: {data.get('updated', '?')}).\n\n"
                    "Continuar donde quedaste?",
                ):
                    self.decisions = data.get("decisions", {})
            except Exception:
                pass

        self.show_resumen()

    # -----------------------------------------------------------------
    # Pantalla 2: resumen del analisis
    # -----------------------------------------------------------------

    def _counts(self):
        auto = sum(1 for r in self.scan_results
                   if not r["ya_lleno"] and r["match"] and r["match"]["tier"] == "AUTO")
        sin_match = sum(1 for r in self.scan_results
                         if not r["ya_lleno"] and r["match"] is None)
        ya_lleno = sum(1 for r in self.scan_results if r["ya_lleno"])
        total_grupos = len(self.review_groups)
        decididos = sum(1 for g in self.review_groups if g["key"] in self.decisions)
        aceptados = sum(1 for v in self.decisions.values() if v == "SI")
        rechazados = sum(1 for v in self.decisions.values() if v == "NO")
        pendientes = total_grupos - decididos
        filas_revisar = sum(len(g["rows"]) for g in self.review_groups)
        return dict(auto=auto, sin_match=sin_match, ya_lleno=ya_lleno,
                    total_grupos=total_grupos, decididos=decididos,
                    aceptados=aceptados, rechazados=rechazados,
                    pendientes=pendientes, filas_revisar=filas_revisar)

    def show_resumen(self):
        self.clear()
        c = self._counts()

        ctk.CTkLabel(self.content, text="Resultado del analisis", font=FONT_H1,
                     text_color=TEXT, anchor="w").pack(anchor="w")
        ctk.CTkLabel(self.content, text=self.osa_path.name if self.osa_path else "",
                     font=FONT_BODY, text_color=MUTED, anchor="w").pack(anchor="w", pady=(2, 20))

        grid = ctk.CTkFrame(self.content, fg_color="transparent")
        grid.pack(fill="x")
        for i in range(4):
            grid.columnconfigure(i, weight=1, uniform="stat")

        self._stat_tile(grid, 0, "✅", c["auto"], "Identificadas\nautomaticamente", SUCCESS, SUCCESS_LIGHT)
        self._stat_tile(grid, 1, "🔍", c["pendientes"], "Por revisar\n(dudosas)", WARNING, WARNING_LIGHT)
        self._stat_tile(grid, 2, "➖", c["sin_match"], "Sin coincidencia\nen tu catalogo", MUTED, MUTED_LIGHT)
        self._stat_tile(grid, 3, "🔒", c["ya_lleno"], "Ya tenian datos\n(no se tocan)", MUTED, MUTED_LIGHT)

        if c["decididos"] > 0:
            info = ctk.CTkFrame(self.content, fg_color=PRIMARY_LIGHT, corner_radius=10)
            info.pack(fill="x", pady=(16, 0))
            ctk.CTkLabel(
                info, text=f"ℹ️  Ya revisaste {c['decididos']} de {c['total_grupos']} casos dudosos "
                           f"({c['aceptados']} aceptados, {c['rechazados']} rechazados) — "
                           f"quedan {c['pendientes']} por decidir.",
                font=FONT_BODY, text_color=PRIMARY, anchor="w", justify="left",
            ).pack(anchor="w", padx=14, pady=10)

        btns = ctk.CTkFrame(self.content, fg_color="transparent")
        btns.pack(fill="x", pady=(26, 0))

        if c["pendientes"] > 0:
            h_button(btns, f"🔍  Revisar coincidencias dudosas  ({c['pendientes']})",
                     self.show_juego, height=48, font=FONT_BTN_BIG).pack(anchor="w", pady=4)
        else:
            done = ctk.CTkFrame(self.content, fg_color=SUCCESS_LIGHT, corner_radius=10)
            done.pack(fill="x", pady=(0, 4))
            ctk.CTkLabel(done, text="🎉  No quedan coincidencias dudosas por revisar.",
                         font=FONT_BODY_BOLD, text_color=SUCCESS_HOVER).pack(anchor="w", padx=14, pady=10)

        h_button(btns, "Generar archivo final para la OSA", self.show_exportar,
                 fg=CARD, hover=MUTED_LIGHT, text_color=TEXT, font=FONT_BODY_BOLD,
                 height=40, border_width=1, border_color=BORDER).pack(anchor="w", pady=6)

        ghost_button(btns, "←  Elegir otros archivos", self.show_inicio).pack(anchor="w", pady=(10, 0))

    def _stat_tile(self, parent, col, icon, number, label, color, tint):
        c = card(parent, fg_color=tint, border_color=tint)
        c.grid(row=0, column=col, sticky="nsew", padx=6)
        ctk.CTkLabel(c, text=icon, font=(F, 18)).pack(pady=(16, 0))
        ctk.CTkLabel(c, text=str(number), font=FONT_STAT_NUM, text_color=color).pack()
        ctk.CTkLabel(c, text=label, font=FONT_SMALL, text_color=TEXT_SECONDARY, justify="center").pack(pady=(0, 16))

    # -----------------------------------------------------------------
    # Pantalla 3: juego de revision
    # -----------------------------------------------------------------

    def show_juego(self):
        self.clear()
        self.queue = deque(i for i, g in enumerate(self.review_groups) if g["key"] not in self.decisions)
        self.history = []

        top = ctk.CTkFrame(self.content, fg_color="transparent")
        top.pack(fill="x")
        ctk.CTkLabel(top, text="¿Es la misma obra?", font=FONT_H1, text_color=TEXT).pack(side="left")
        ghost_button(top, "Terminar por ahora  →  exportar", self.show_exportar).pack(side="right")

        self.progress_label = ctk.CTkLabel(self.content, font=FONT_BODY, text_color=MUTED, anchor="w")
        self.progress_label.pack(anchor="w", pady=(10, 4))
        self.progress_bar = ctk.CTkProgressBar(self.content, height=10, corner_radius=6,
                                                progress_color=PRIMARY, fg_color=MUTED_LIGHT)
        self.progress_bar.pack(fill="x", pady=(0, 18))

        cards = ctk.CTkFrame(self.content, fg_color="transparent")
        cards.pack(fill="x")
        cards.columnconfigure(0, weight=1, uniform="cards")
        cards.columnconfigure(1, weight=1, uniform="cards")

        osa_box = card(cards)
        osa_box.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        bd_box = card(cards)
        bd_box.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        osa_inner = ctk.CTkFrame(osa_box, fg_color="transparent")
        osa_inner.pack(fill="both", expand=True, padx=20, pady=18)
        osa_head = ctk.CTkFrame(osa_inner, fg_color="transparent")
        osa_head.pack(anchor="w")
        if self.logo_osa_card:
            ctk.CTkLabel(osa_head, text="", image=self.logo_osa_card).pack(side="left", padx=(0, 10))
        badge(osa_head, "REPORTA LA OSA", "#1D4ED8", "#DBEAFE").pack(side="left")
        self.lbl_osa_titulo = ctk.CTkLabel(osa_inner, font=FONT_CARD_TITLE, text_color=TEXT,
                                            anchor="w", justify="left", wraplength=400)
        self.lbl_osa_titulo.pack(anchor="w", pady=(12, 14))
        self.lbl_osa_autor = ctk.CTkLabel(osa_inner, font=FONT_BODY, text_color=TEXT_SECONDARY,
                                           anchor="w", justify="left", wraplength=400)
        self.lbl_osa_autor.pack(anchor="w", pady=2)
        self.lbl_osa_interprete = ctk.CTkLabel(osa_inner, font=FONT_BODY, text_color=MUTED,
                                                anchor="w", justify="left", wraplength=400)
        self.lbl_osa_interprete.pack(anchor="w", pady=2)
        self.lbl_osa_filas = ctk.CTkLabel(osa_inner, font=FONT_SMALL, text_color=MUTED,
                                           fg_color=MUTED_LIGHT, corner_radius=8, anchor="w")
        self.lbl_osa_filas.pack(anchor="w", pady=(14, 0), ipadx=4, ipady=2)

        bd_inner = ctk.CTkFrame(bd_box, fg_color="transparent")
        bd_inner.pack(fill="both", expand=True, padx=20, pady=18)
        bd_head = ctk.CTkFrame(bd_inner, fg_color="transparent")
        bd_head.pack(anchor="w")
        if self.logo_fuentes_card:
            logo_box = ctk.CTkFrame(bd_head, fg_color=CARD, corner_radius=8, border_width=1, border_color=BORDER)
            logo_box.pack(side="left", padx=(0, 10))
            ctk.CTkLabel(logo_box, text="", image=self.logo_fuentes_card).pack(padx=4, pady=4)
        badge(bd_head, "BASE EDIMUSICA", PRIMARY, PRIMARY_LIGHT).pack(side="left")
        self.lbl_bd_titulo = ctk.CTkLabel(bd_inner, font=FONT_CARD_TITLE, text_color=TEXT,
                                           anchor="w", justify="left", wraplength=400)
        self.lbl_bd_titulo.pack(anchor="w", pady=(12, 14))
        self.lbl_bd_autor = ctk.CTkLabel(bd_inner, font=FONT_BODY, text_color=TEXT_SECONDARY,
                                          anchor="w", justify="left", wraplength=400)
        self.lbl_bd_autor.pack(anchor="w", pady=2)
        self.lbl_bd_pct = ctk.CTkLabel(bd_inner, font=FONT_BODY_BOLD, text_color=SUCCESS_HOVER, anchor="w")
        self.lbl_bd_pct.pack(anchor="w", pady=(10, 2))
        self.lbl_bd_score = ctk.CTkLabel(bd_inner, font=FONT_SMALL, text_color=MUTED,
                                          anchor="w", justify="left", wraplength=400)
        self.lbl_bd_score.pack(anchor="w", pady=(10, 0))

        btns = ctk.CTkFrame(self.content, fg_color="transparent")
        btns.pack(fill="x", pady=(22, 6))
        btns.columnconfigure(0, weight=1, uniform="act")
        btns.columnconfigure(1, weight=1, uniform="act")
        btns.columnconfigure(2, weight=1, uniform="act")

        h_button(btns, "✕   No coincide   (←)", self.decide_no, fg=DANGER, hover=DANGER_HOVER,
                 height=58, font=FONT_BTN_BIG, corner_radius=14).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        h_button(btns, "⏭   Despues   (espacio)", self.decide_skip, fg=CARD, hover=MUTED_LIGHT,
                 text_color=TEXT, height=58, font=FONT_BTN_BIG, corner_radius=14,
                 border_width=1, border_color=BORDER).grid(row=0, column=1, sticky="ew", padx=6)
        h_button(btns, "✓   Si, es la misma   (→)", self.decide_yes, fg=SUCCESS, hover=SUCCESS_HOVER,
                 height=58, font=FONT_BTN_BIG, corner_radius=14).grid(row=0, column=2, sticky="ew", padx=(6, 0))

        bottom = ctk.CTkFrame(self.content, fg_color="transparent")
        bottom.pack(fill="x")
        ghost_button(bottom, "↩  Deshacer ultima decision  (Ctrl+Z)", self.undo).pack(anchor="w")

        self.bind("<Left>", lambda e: self.decide_no())
        self.bind("<Right>", lambda e: self.decide_yes())
        self.bind("<Return>", lambda e: self.decide_yes())
        self.bind("<space>", lambda e: self.decide_skip())
        self.bind("<Control-z>", lambda e: self.undo())

        self._refresh_card()

    def _refresh_card(self):
        total = len(self.review_groups)
        decididos = total - len(self.queue)
        self.progress_label.configure(
            text=f"{decididos} de {total} decisiones tomadas   ·   "
                 f"{sum(1 for v in self.decisions.values() if v=='SI')} aceptadas   ·   "
                 f"{sum(1 for v in self.decisions.values() if v=='NO')} rechazadas"
        )
        self.progress_bar.set(decididos / total if total else 0)

        if not self.queue:
            self.show_resumen()
            return

        idx = self.queue[0]
        g = self.review_groups[idx]
        sample = g["sample"]
        match = g["match"]
        summary = match["summary"]

        self.lbl_osa_titulo.configure(text=sample["titulo"])
        self.lbl_osa_autor.configure(text=f"✍️  {sample['autor'] or '(sin dato)'}")
        self.lbl_osa_interprete.configure(text=f"🎤  {sample['interprete'] or '(sin dato)'}")
        n_rows = len(g["rows"])
        self.lbl_osa_filas.configure(
            text=f"  Aplica a {n_rows} filas de la consulta OSA  " if n_rows > 1
            else "  Aplica a 1 fila de la consulta OSA  "
        )

        self.lbl_bd_titulo.configure(text=summary["titulo"])
        self.lbl_bd_autor.configure(text="✍️  " + mc.format_authors(summary["authors"]))
        self.lbl_bd_pct.configure(text=f"💰  Edimusica administraria: {summary['total_pct']:.2f}%")
        self.lbl_bd_score.configure(
            text=f"Tipo de match: {match['match_type']}   ·   score titulo: {match['title_score']}   ·   "
                 f"score autor: {match['author_score']}   ·   codigo: {summary['codant']}"
        )

    def _decide(self, value):
        if not self.queue:
            return
        idx = self.queue.popleft()
        key = self.review_groups[idx]["key"]
        self.decisions[key] = value
        self.history.append(idx)
        self.save_session()
        self._refresh_card()

    def decide_yes(self):
        self._decide("SI")

    def decide_no(self):
        self._decide("NO")

    def decide_skip(self):
        if not self.queue:
            return
        idx = self.queue.popleft()
        if self.queue:
            self.queue.append(idx)
            self._refresh_card()
        else:
            # era el unico pendiente: no tiene caso re-mostrarlo en bucle
            self.queue.append(idx)
            messagebox.showinfo(
                "Solo queda este caso",
                "Es el unico caso pendiente. Decide Si/No o ve a exportar dejandolo pendiente.",
            )

    def undo(self):
        if not self.history:
            return
        idx = self.history.pop()
        key = self.review_groups[idx]["key"]
        self.decisions.pop(key, None)
        self.queue.appendleft(idx)
        self.save_session()
        self._refresh_card()

    # -----------------------------------------------------------------
    # Pantalla 4: exportar
    # -----------------------------------------------------------------

    def show_exportar(self):
        self.clear()
        for seq in ("<Left>", "<Right>", "<Return>", "<space>", "<Control-z>"):
            self.unbind(seq)

        c = self._counts()
        ctk.CTkLabel(self.content, text="Generar archivo final", font=FONT_H1,
                     text_color=TEXT, anchor="w").pack(anchor="w")

        info_card = card(self.content)
        info_card.pack(fill="x", pady=(18, 0))
        ctk.CTkLabel(
            info_card,
            text="Se creara una COPIA del archivo de la OSA (el original no se modifica),\n"
                 "llenando EDITOR y % solo en las obras identificadas automaticamente o\n"
                 "confirmadas por ti. Lo pendiente o rechazado no se toca.",
            font=FONT_BODY, text_color=TEXT_SECONDARY, justify="left", anchor="w",
        ).pack(anchor="w", padx=18, pady=16)

        chips = ctk.CTkFrame(self.content, fg_color="transparent")
        chips.pack(fill="x", pady=(14, 0))
        badge(chips, f"✅  Se llenaran: {c['auto'] + c['aceptados']}", SUCCESS_HOVER, SUCCESS_LIGHT).pack(side="left")
        quedan_pendientes = max(c["total_grupos"] - c["aceptados"] - c["rechazados"], 0)
        badge(chips, f"🔍  Quedaran pendientes: {quedan_pendientes}", WARNING, WARNING_LIGHT).pack(side="left", padx=10)

        self.export_status = StringVar(value="")
        ctk.CTkLabel(self.content, textvariable=self.export_status, font=FONT_BODY,
                     text_color=MUTED, anchor="w").pack(anchor="w", pady=(20, 4))

        btns = ctk.CTkFrame(self.content, fg_color="transparent")
        btns.pack(fill="x", pady=10)
        self.btn_exportar = h_button(btns, "Generar archivo para la OSA", self._start_export,
                                      height=48, font=FONT_BTN_BIG)
        self.btn_exportar.pack(side="left")
        ghost_button(btns, "←  Volver a revisar", self.show_resumen).pack(side="left", padx=16)

    def _start_export(self):
        self.btn_exportar.configure(state="disabled")
        self._start_dots("Generando archivo", self.export_status)
        threading.Thread(target=self._run_export, daemon=True).start()

    def _run_export(self):
        try:
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            output_path = self.osa_path.with_name(f"{self.osa_path.stem}_EDIMUSICA_{timestamp}.xlsx")
            log_path = self.osa_path.with_name(f"log_matching_{timestamp}.xlsx")

            wb2 = openpyxl.load_workbook(self.osa_path, data_only=False)
            ws2 = wb2[mc.OSA_SHEET]
            counters, log_rows = mc.apply_results(ws2, self.scan_results, self.decisions)

            log_wb = openpyxl.Workbook()
            log_ws = log_wb.active
            log_ws.title = "log"
            log_ws.append([
                "Fila OSA", "Titulo OSA", "Autor OSA", "Decision", "Tipo match",
                "Score titulo", "Score autor", "Titulo BD", "Autor(es) BD",
                "% total BD", "Codigo BD",
            ])
            for row in log_rows:
                log_ws.append(list(row))

            wb2.save(output_path)
            log_wb.save(log_path)
        except Exception as exc:  # noqa: BLE001
            self.after(0, self._export_failed, str(exc))
            return
        self.after(0, self._export_done, output_path, log_path, counters)

    def _export_failed(self, message):
        self._stop_dots()
        self.btn_exportar.configure(state="normal")
        self.export_status.set("")
        messagebox.showerror("Error al generar el archivo", message)

    def _export_done(self, output_path, log_path, counters):
        self._stop_dots()
        self.clear()

        ctk.CTkLabel(self.content, text="✅  ¡Archivo generado!", font=FONT_H1,
                     text_color=SUCCESS_HOVER, anchor="w").pack(anchor="w", pady=(0, 18))

        paths_card = card(self.content)
        paths_card.pack(fill="x")
        inner = ctk.CTkFrame(paths_card, fg_color="transparent")
        inner.pack(fill="x", padx=18, pady=16)
        ctk.CTkLabel(inner, text="Archivo para enviar a la OSA", font=FONT_SMALL, text_color=MUTED,
                     anchor="w").pack(anchor="w")
        ctk.CTkLabel(inner, text=str(output_path), font=FONT_MONO, text_color=TEXT,
                     anchor="w", wraplength=980, justify="left").pack(anchor="w", pady=(2, 12))
        ctk.CTkLabel(inner, text="Log de auditoria", font=FONT_SMALL, text_color=MUTED,
                     anchor="w").pack(anchor="w")
        ctk.CTkLabel(inner, text=str(log_path), font=FONT_MONO, text_color=TEXT,
                     anchor="w", wraplength=980, justify="left").pack(anchor="w", pady=(2, 0))

        chips = ctk.CTkFrame(self.content, fg_color="transparent")
        chips.pack(fill="x", pady=(18, 0))
        badge(chips, f"Automaticas: {counters['auto']}", SUCCESS_HOVER, SUCCESS_LIGHT).pack(side="left")
        badge(chips, f"Confirmadas: {counters['confirmado']}", SUCCESS_HOVER, SUCCESS_LIGHT).pack(side="left", padx=8)
        badge(chips, f"Rechazadas: {counters['rechazado']}", DANGER, DANGER_LIGHT).pack(side="left")
        badge(chips, f"Pendientes: {counters['pendiente']}", WARNING, WARNING_LIGHT).pack(side="left", padx=8)
        badge(chips, f"Sin coincidencia: {counters['sin_match']}", MUTED, MUTED_LIGHT).pack(side="left")

        btns = ctk.CTkFrame(self.content, fg_color="transparent")
        btns.pack(fill="x", pady=26)
        h_button(btns, "📂  Abrir carpeta", lambda: webbrowser.open(str(output_path.parent)),
                 height=42).pack(side="left")
        h_button(btns, "Volver al resumen", self.show_resumen, fg=CARD, hover=MUTED_LIGHT,
                 text_color=TEXT, height=42, border_width=1, border_color=BORDER).pack(side="left", padx=10)
        ghost_button(btns, "Empezar con otro archivo", self.show_inicio).pack(side="left")


if __name__ == "__main__":
    app = App()
    app.mainloop()
