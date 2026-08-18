"""
Motor de cruce entre la CONSULTA OSA y el catalogo de obras de Edimusica.

Modulo puro (sin I/O de consola ni argparse) para que lo puedan usar tanto
el script de linea de comandos (recon_osa.py) como la interfaz grafica
(recon_osa_gui.py).
"""

import re
import unicodedata
from collections import defaultdict

import openpyxl
from rapidfuzz import fuzz, process

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------

EDITOR_NAME = "EDIMUSICA"

OSA_SHEET = "Hoja1"
OSA_HEADER_ROW = 2
OSA_FIRST_DATA_ROW = 3

COL_TITULO = 4          # D - TITULO DE LA OBRA
COL_AUTOR = 5           # E - AUTOR
COL_INTERPRETE = 6      # F - INTERPRETE
COL_EDITOR = 9          # I - EDITOR
COL_PCT = 10            # J - %
COL_COMENTARIOS = 11    # K - COMENTARIOS

DB_SHEET = "Sheet1"

# --- Umbrales de decision -----------------------------------------------
# Criterio general (ver docstring de match_title_author): ante la duda, NO
# sugerir. Mejor una obra sin identificar que una coincidencia enganosa.
FUZZY_CANDIDATE_CUTOFF = 88      # score minimo de titulo para considerar candidato
TITLE_WORD_OVERLAP_MIN = 0.75    # fraccion de palabras del titulo corto que deben aparecer en el largo
TIER1_FUZZY_TITLE_MIN = 95       # score de titulo (no exacto) para auto-llenar
AUTHOR_CONFIRM_MIN = 85          # similitud de autor para confirmar desambiguacion (AUTO)
AUTHOR_MARGIN_MIN = 12           # margen minimo entre mejor y segundo candidato
REVISAR_AUTHOR_FLOOR = 50        # por debajo de esto, ni se sugiere para revisar
AUTHOR_FLOOR_PENALTY_PER_COAUTHOR = 15  # sube el piso por cada coautor extra a comparar
MAX_DOC_FREQ_FOR_BLOCKING = 450  # palabras muy comunes no sirven para bloqueo
MIN_WORD_LEN_FOR_BLOCKING = 3

STOPWORDS = {
    "DE", "LA", "EL", "Y", "EN", "CON", "A", "LOS", "LAS", "QUE", "MI", "TU",
    "SE", "SU", "NO", "UN", "UNA", "POR", "PARA", "DEL", "AL", "LO", "ES",
    "TE", "SI", "ME", "MAS", "O", "E", "TODO", "TODA",
}

UNIDENTIFIED_VALUES = {
    "", "NO IDENTIFICADO", "SIN IDENTIFICAR", "DESCONOCIDO",
    "AUTOR NO IDENTIFICADO", "SIN AUTOR", "N A", "NA",
}


# ---------------------------------------------------------------------------
# Normalizacion de texto
# ---------------------------------------------------------------------------

def normalize(s):
    if s is None:
        return ""
    s = str(s).upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\(\s*\d+\s*\)", " ", s)   # quita anotaciones tipo "(100)"
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def split_authors(raw):
    """Separa varios coautores de un mismo campo AUTOR.

    Solo se parte en separadores inequivocos de coautoria ("/", ";", " Y ",
    "&"). La coma NO se usa como separador: en los datos de la OSA aparece
    mayoritariamente en formato "Apellido, Nombre" de un unico autor (igual
    a como Edimusica guarda sus propios nombres una vez se quita la coma),
    y partir por coma rompe ese caso frecuente sin resolver de forma fiable
    los casos reales de lista de autores separados por coma.
    """
    if raw is None:
        return []
    text = str(raw)
    parts = re.split(r"\s*/\s*|\s*;\s*|\s+Y\s+|\s+&\s+", text, flags=re.IGNORECASE)
    return [p.strip() for p in parts if p and p.strip()]


def is_unidentified(raw):
    return normalize(raw) in UNIDENTIFIED_VALUES


# ---------------------------------------------------------------------------
# Carga y indices de la base de Edimusica
# ---------------------------------------------------------------------------

def load_db(db_path):
    wb = openpyxl.load_workbook(db_path, read_only=True, data_only=True)
    ws = wb[DB_SHEET]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None:
            continue
        titulo, nomautor, codigo, codant, pct = (list(r) + [None] * 5)[:5]
        if not titulo:
            continue
        try:
            pctval = float(str(pct).replace(",", "."))
        except (TypeError, ValueError):
            pctval = 0.0
        rows.append({
            "titulo": titulo,
            "titulo_norm": normalize(titulo),
            "autor": nomautor,
            "autor_norm": normalize(nomautor),
            "codigo": codigo,
            "codant": codant,
            "pct": pctval,
        })
    wb.close()
    return rows


def build_exact_index(rows):
    idx = defaultdict(list)
    for i, r in enumerate(rows):
        idx[r["titulo_norm"]].append(i)
    return idx


def build_word_index(rows):
    doc_freq = defaultdict(int)
    word_rows = defaultdict(set)
    for i, r in enumerate(rows):
        words = {w for w in r["titulo_norm"].split()
                 if len(w) >= MIN_WORD_LEN_FOR_BLOCKING and w not in STOPWORDS}
        for w in words:
            doc_freq[w] += 1
            word_rows[w].add(i)
    return {w: idxs for w, idxs in word_rows.items() if doc_freq[w] <= MAX_DOC_FREQ_FOR_BLOCKING}


def get_fuzzy_candidates(title_norm, blocking_index, limit_pool=800):
    words = [w for w in title_norm.split()
             if len(w) >= MIN_WORD_LEN_FOR_BLOCKING and w not in STOPWORDS]
    candidate_ids = set()
    for w in words:
        if w in blocking_index:
            candidate_ids |= blocking_index[w]
            if len(candidate_ids) > limit_pool:
                break
    return candidate_ids


class MatchIndex:
    """Agrupa la base de Edimusica cargada junto con sus indices de busqueda."""

    def __init__(self, db_path):
        self.rows = load_db(db_path)
        self.exact_index = build_exact_index(self.rows)
        self.blocking_index = build_word_index(self.rows)


# ---------------------------------------------------------------------------
# Logica de matching
# ---------------------------------------------------------------------------

def summarize_group(rows, indices):
    total_pct = round(sum(rows[i]["pct"] for i in indices), 2)
    authors = [(rows[i]["autor"], rows[i]["pct"]) for i in indices]
    titulo = rows[indices[0]]["titulo"]
    codant = rows[indices[0]]["codant"]
    return {"total_pct": total_pct, "authors": authors, "titulo": titulo, "codant": codant}


def best_group_by_author(rows, candidate_indices, osa_authors_norm):
    by_group = defaultdict(list)
    for i in candidate_indices:
        by_group[rows[i]["codant"]].append(i)

    scored = []
    for codant, idxs in by_group.items():
        best_author_score = 0
        if osa_authors_norm:
            for i in idxs:
                a_norm = rows[i]["autor_norm"]
                for oa in osa_authors_norm:
                    sc = fuzz.token_sort_ratio(a_norm, oa)
                    if sc > best_author_score:
                        best_author_score = sc
        scored.append((best_author_score, codant, idxs))
    scored.sort(key=lambda x: -x[0])
    return scored


def significant_words(title_norm):
    return {w for w in title_norm.split() if len(w) >= MIN_WORD_LEN_FOR_BLOCKING and w not in STOPWORDS}


def word_overlap_ratio(words_a, words_b):
    """Fraccion de las palabras del titulo mas corto que aparecen (exactas)
    en el titulo mas largo. Sirve para descartar coincidencias que solo
    comparten una palabra suelta (p.ej. "DOS FLORES" vs "LOS DOS") aunque
    el score de caracteres (Levenshtein) las haga ver parecidas: dos
    titulos cortos comparten muchas letras por azar.
    """
    if not words_a or not words_b:
        return 0.0
    shorter, longer = (words_a, words_b) if len(words_a) <= len(words_b) else (words_b, words_a)
    matched = sum(1 for w in shorter if w in longer)
    return matched / len(shorter)


def title_is_distinctive(title_norm):
    """Un titulo corto/generico ("LOCO", "AZUL", "SIN TI", "TE QUIERO") es
    demasiado comun para identificar una obra sin una señal de autor fuerte:
    en un catalogo de ~20.000 obras es muy probable que existan varias
    canciones distintas con el mismo titulo generico. Se usa el largo total
    de la frase (no solo las palabras "significativas" sin stopwords, que
    dejaba pasar frases cortas como "EN MI X" con una sola palabra de
    contenido) como proxy de que tan especifico es el titulo.
    """
    total_words = len(title_norm.split())
    if total_words >= 3:
        return True
    if total_words == 2:
        return len(title_norm) >= 10
    return len(title_norm) >= 12


def effective_author_floor(title_norm, group_size):
    """Piso minimo de score de autor para siquiera sugerir un candidato
    como REVISAR (por debajo, se descarta por completo).

    Sube en dos casos, y toma el mas exigente de los dos (no se suman,
    para no volverlo imposible de alcanzar):
      - titulo generico: sin un titulo distintivo, un titulo exacto no es
        evidencia fuerte por si solo (hay homonimos), asi que se exige el
        mismo nivel de autor que para auto-llenar.
      - obra con varios coautores: al comparar el autor de la OSA contra
        cada coautor y quedarse con el mejor resultado, entre mas
        coautores tenga la obra mas probable es que UNO de ellos parezca
        coincidir por pura casualidad (apellido comun, nombre compuesto
        compartido) aunque no sea la misma obra. Se compenso subiendo el
        piso segun cuantos coautores hay que comparar.
    """
    floor = REVISAR_AUTHOR_FLOOR
    if not title_is_distinctive(title_norm):
        floor = max(floor, AUTHOR_CONFIRM_MIN)
    if group_size > 1:
        floor = max(floor, REVISAR_AUTHOR_FLOOR + AUTHOR_FLOOR_PENALTY_PER_COAUTHOR * (group_size - 1))
    return min(floor, 90)


def match_title_author(match_index, titulo_raw, autor_raw):
    """Devuelve el mejor match (dict) o None si no hay ningun candidato lo
    bastante confiable como para sugerirlo siquiera.

    Principio rector (pedido explicitamente: "razonamiento muy critico"):
    ante la duda, NO sugerir. Es preferible dejar una obra sin identificar
    que sugerir -o peor, auto-llenar- una coincidencia que solo comparte
    una palabra de titulo o un apellido comun. Por eso hay dos filtros que
    se aplican SIEMPRE, incluso antes de pensar en "REVISAR":
      - titulo: si no es exacto, se exige solapamiento real de palabras
        (no solo similitud de caracteres).
      - autor: si el autor de la OSA es conocido, un score de autor muy
        bajo descarta el candidato por completo (no se sugiere ni para
        revisar) aunque el titulo sea exacto - un titulo exacto pero con
        un autor sin ninguna relacion suele ser una obra homonima
        distinta, no la misma obra.
    """
    rows = match_index.rows
    title_norm = normalize(titulo_raw)
    if not title_norm:
        return None

    unidentified = is_unidentified(autor_raw)
    # Si el autor OSA es "No identificado" no hay ninguna señal real de
    # autor: se deja la lista vacia en vez de comparar el texto literal
    # "no identificado" contra los autores de la base (eso producia un
    # "score de autor" bajo pero completamente ficticio, engañoso al
    # mostrarlo en la revision).
    osa_authors_norm = [] if unidentified else [
        a for a in (normalize(x) for x in split_authors(autor_raw)) if a
    ]

    exact_candidates = match_index.exact_index.get(title_norm, [])

    if exact_candidates:
        scored_groups = best_group_by_author(rows, exact_candidates, osa_authors_norm)
        top_score, _, top_idxs = scored_groups[0]
        second_score = scored_groups[1][0] if len(scored_groups) > 1 else -1

        if unidentified:
            # Sin autor de referencia, la unica evidencia es el titulo: solo
            # vale la pena sugerirlo si es exacto, inequivoco (una sola obra
            # en toda la base con ese titulo) y no es un titulo generico.
            if len(scored_groups) == 1 and title_is_distinctive(title_norm):
                tier = "REVISAR"
            else:
                return None
        elif top_score >= AUTHOR_CONFIRM_MIN and (top_score - second_score) >= AUTHOR_MARGIN_MIN:
            tier = "AUTO"
        elif top_score >= effective_author_floor(title_norm, len(top_idxs)):
            tier = "REVISAR"
        else:
            return None

        return {
            "tier": tier,
            "title_score": 100,
            "author_score": "N/A" if unidentified else round(top_score, 1),
            "summary": summarize_group(rows, top_idxs),
            "match_type": "EXACTO",
        }

    candidate_ids = get_fuzzy_candidates(title_norm, match_index.blocking_index)
    if not candidate_ids:
        return None

    choices = {i: rows[i]["titulo_norm"] for i in candidate_ids}
    results = process.extract(
        title_norm, choices, scorer=fuzz.token_sort_ratio,
        limit=15, score_cutoff=FUZZY_CANDIDATE_CUTOFF,
    )
    if not results:
        return None

    # Filtro de solapamiento de palabras: descarta candidatos que solo
    # deben su score de caracteres a compartir una palabra suelta.
    query_words = significant_words(title_norm)
    results = [
        r for r in results
        if word_overlap_ratio(query_words, significant_words(rows[r[2]]["titulo_norm"])) >= TITLE_WORD_OVERLAP_MIN
    ]
    if not results:
        return None

    matched_ids = [key for _, _, key in results]
    best_title_score = results[0][1]

    scored_groups = best_group_by_author(rows, matched_ids, osa_authors_norm)
    top_score, _, top_idxs = scored_groups[0]
    second_score = scored_groups[1][0] if len(scored_groups) > 1 else -1

    if unidentified:
        # Titulo aproximado (no exacto) + sin autor de referencia es la
        # combinacion mas riesgosa: se exige el titulo casi identico,
        # inequivoco y distintivo antes de sugerirlo.
        if (best_title_score >= TIER1_FUZZY_TITLE_MIN and len(scored_groups) == 1
                and title_is_distinctive(title_norm)):
            tier = "REVISAR"
        else:
            return None
    elif (best_title_score >= TIER1_FUZZY_TITLE_MIN and top_score >= AUTHOR_CONFIRM_MIN
            and (top_score - second_score) >= AUTHOR_MARGIN_MIN):
        tier = "AUTO"
    elif top_score >= effective_author_floor(title_norm, len(top_idxs)):
        tier = "REVISAR"
    else:
        return None

    return {
        "tier": tier,
        "title_score": round(best_title_score, 1),
        "author_score": "N/A" if unidentified else round(top_score, 1),
        "summary": summarize_group(rows, top_idxs),
        "match_type": "APROXIMADO",
    }


# ---------------------------------------------------------------------------
# Formato de comentarios
# ---------------------------------------------------------------------------

def format_authors(authors):
    return ", ".join(f"{a} ({p:.2f}%)" for a, p in authors)


def build_comment_auto(unidentified, summary):
    if not unidentified:
        return None
    return f"AUTOR(ES) IDENTIFICADO(S) SEGUN BD EDIMUSICA: {format_authors(summary['authors'])}"


def build_comment_review(match):
    s = match["summary"]
    return (
        f"REVISAR POSIBLE COINCIDENCIA EDIMUSICA ({match['match_type']}, "
        f"score titulo={match['title_score']}, score autor={match['author_score']}): "
        f"Titulo BD='{s['titulo']}' | Autor(es) BD: {format_authors(s['authors'])} "
        f"| % total BD={s['total_pct']:.2f} | Codigo={s['codant']}"
    )


# ---------------------------------------------------------------------------
# Escaneo completo de un archivo OSA (usado por la CLI y como base de la GUI)
# ---------------------------------------------------------------------------

def scan_osa_rows(ws, match_index, first_row=None, last_row=None):
    """Recorre las filas de datos de la hoja OSA y calcula el match de cada una.

    No escribe nada en la hoja. Devuelve una lista de dicts, uno por fila con
    titulo, uno de:
      {"row": int, "titulo": str, "autor": str, "match": dict|None,
       "ya_lleno": bool}
    """
    first_row = first_row or OSA_FIRST_DATA_ROW
    last_row = last_row or ws.max_row

    results = []
    for row_idx in range(first_row, last_row + 1):
        titulo_raw = ws.cell(row=row_idx, column=COL_TITULO).value
        autor_raw = ws.cell(row=row_idx, column=COL_AUTOR).value
        interprete_raw = ws.cell(row=row_idx, column=COL_INTERPRETE).value

        if not titulo_raw:
            continue

        editor_val = ws.cell(row=row_idx, column=COL_EDITOR).value
        pct_val = ws.cell(row=row_idx, column=COL_PCT).value
        ya_lleno = editor_val not in (None, "") or pct_val not in (None, "")

        match = None if ya_lleno else match_title_author(match_index, titulo_raw, autor_raw)

        results.append({
            "row": row_idx,
            "titulo": titulo_raw,
            "autor": autor_raw,
            "interprete": interprete_raw,
            "match": match,
            "ya_lleno": ya_lleno,
        })
    return results


def review_key(titulo_raw, autor_raw):
    return normalize(titulo_raw) + "||" + normalize(autor_raw)


def build_review_groups(scan_results):
    """Agrupa las filas REVISAR por (titulo, autor) normalizado.

    Varias filas de la OSA pueden reportar la misma obra (distintos
    interpretes); comparten titulo+autor => mismo resultado de matching
    => se revisan como una sola decision que luego se aplica a todas.
    """
    groups = {}
    order = []
    for item in scan_results:
        m = item["match"]
        if item["ya_lleno"] or m is None or m["tier"] != "REVISAR":
            continue
        key = review_key(item["titulo"], item["autor"])
        if key not in groups:
            groups[key] = {"key": key, "rows": [], "sample": item, "match": m}
            order.append(key)
        groups[key]["rows"].append(item["row"])
    return [groups[k] for k in order]


# ---------------------------------------------------------------------------
# Aplicar resultados + decisiones humanas a la hoja de la OSA
# ---------------------------------------------------------------------------

def apply_results(ws, scan_results, decisions):
    """Escribe EDITOR/%/COMENTARIOS en `ws` segun el tier de cada fila y las
    decisiones humanas (para las REVISAR). No toca filas ya llenas ni sin
    match. Devuelve (contadores dict, log_rows list-of-tuples).
    """
    counters = {"auto": 0, "confirmado": 0, "rechazado": 0, "pendiente": 0, "sin_match": 0, "ya_lleno": 0}
    log_rows = []

    for item in scan_results:
        if item["ya_lleno"]:
            counters["ya_lleno"] += 1
            continue

        match = item["match"]
        if match is None:
            counters["sin_match"] += 1
            continue

        row_idx = item["row"]
        unidentified = is_unidentified(item["autor"])
        summary = match["summary"]

        editor_cell = ws.cell(row=row_idx, column=COL_EDITOR)
        pct_cell = ws.cell(row=row_idx, column=COL_PCT)
        comentarios_cell = ws.cell(row=row_idx, column=COL_COMENTARIOS)

        if match["tier"] == "AUTO":
            decision = "AUTO"
            counters["auto"] += 1
        else:
            key = review_key(item["titulo"], item["autor"])
            dec = decisions.get(key)
            if dec == "SI":
                decision = "CONFIRMADO_MANUAL"
                counters["confirmado"] += 1
            elif dec == "NO":
                decision = "RECHAZADO_MANUAL"
                counters["rechazado"] += 1
            else:
                decision = "PENDIENTE"
                counters["pendiente"] += 1

        if decision in ("AUTO", "CONFIRMADO_MANUAL"):
            editor_cell.value = EDITOR_NAME
            pct_cell.value = summary["total_pct"]
            if comentarios_cell.value in (None, ""):
                comment = build_comment_auto(unidentified, summary)
                if comment:
                    comentarios_cell.value = comment
        elif decision == "PENDIENTE":
            if comentarios_cell.value in (None, ""):
                comentarios_cell.value = build_comment_review(match)
        # RECHAZADO_MANUAL: no se toca nada.

        log_rows.append((
            row_idx, item["titulo"], item["autor"], decision, match["match_type"],
            match["title_score"], match["author_score"], summary["titulo"],
            format_authors(summary["authors"]), summary["total_pct"], summary["codant"],
        ))

    return counters, log_rows
