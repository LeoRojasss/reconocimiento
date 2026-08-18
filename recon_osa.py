"""
Cruce de la consulta de la OSA contra el catalogo de obras de Edimusica.

Lee "CONSULTA OSA_1Q 2026.xlsx" (formato que envia la OSA) y
"Obras_edimusica.xlsx" (nuestra base de obras), identifica que obras de la
consulta administramos, y genera una COPIA del archivo de la OSA con las
columnas EDITOR, % y COMENTARIOS llenas para los casos de alta confianza.
Los casos dudosos NO se llenan: se anotan en COMENTARIOS como sugerencia
para revision manual. El archivo original de la OSA nunca se modifica.

Uso:
    python recon_osa.py [--limit N] [--osa RUTA] [--db RUTA]

    --limit N   Procesa solo las primeras N filas de datos (pruebas rapidas).
"""

import argparse
import re
import time
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl
from rapidfuzz import fuzz, process

# ---------------------------------------------------------------------------
# Configuracion
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
OSA_FILE_DEFAULT = BASE_DIR / "CONSULTA OSA_1Q 2026.xlsx"
DB_FILE_DEFAULT = BASE_DIR / "Obras_edimusica.xlsx"

EDITOR_NAME = "EDIMUSICA"

OSA_SHEET = "Hoja1"
OSA_HEADER_ROW = 2
OSA_FIRST_DATA_ROW = 3

COL_TITULO = 4        # D - TITULO DE LA OBRA
COL_AUTOR = 5          # E - AUTOR
COL_EDITOR = 9         # I - EDITOR
COL_PCT = 10            # J - %
COL_COMENTARIOS = 11    # K - COMENTARIOS

DB_SHEET = "Sheet1"

# --- Umbrales de decision -----------------------------------------------
FUZZY_CANDIDATE_CUTOFF = 82   # score minimo de titulo para considerar candidato
TIER1_FUZZY_TITLE_MIN = 95     # score de titulo (no exacto) para auto-llenar
AUTHOR_CONFIRM_MIN = 85        # similitud de autor para confirmar desambiguacion
AUTHOR_MARGIN_MIN = 12         # margen minimo entre mejor y segundo candidato
MAX_DOC_FREQ_FOR_BLOCKING = 450  # palabras muy comunes no sirven para bloqueo
MIN_WORD_LEN_FOR_BLOCKING = 3

STOPWORDS = {
    "DE", "LA", "EL", "Y", "EN", "CON", "A", "LOS", "LAS", "QUE", "MI", "TU",
    "SE", "SU", "NO", "UN", "UNA", "POR", "PARA", "DEL", "AL", "LO", "ES",
    "TE", "SI", "ME", "MAS", "O", "E", "TODO", "TODA",
}

UNIDENTIFIED_VALUES = {
    "", "NO IDENTIFICADO", "SIN IDENTIFICAR", "DESCONOCIDO",
    "AUTOR NO IDENTIFICADO", "SIN AUTOR", "N A", "NA",
}


# ---------------------------------------------------------------------------
# Normalizacion de texto
# ---------------------------------------------------------------------------

def normalize(s):
    if s is None:
        return ""
    s = str(s).upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\(\s*\d+\s*\)", " ", s)   # quita anotaciones tipo "(100)"
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def split_authors(raw):
    """Separa varios coautores de un mismo campo AUTOR.

    Solo se parte en separadores inequivocos de coautoria ("/", ";", " Y ",
    "&"). La coma NO se usa como separador: en los datos de la OSA aparece
    mayoritariamente en formato "Apellido, Nombre" de un unico autor (igual
    a como Edimusica guarda sus propios nombres una vez se quita la coma),
    y partir por coma rompe ese caso frecuente sin resolver de forma fiable
    los casos reales de lista de autores separados por coma.
    """
    if raw is None:
        return []
    text = str(raw)
    parts = re.split(r"\s*/\s*|\s*;\s*|\s+Y\s+|\s+&\s+", text, flags=re.IGNORECASE)
    return [p.strip() for p in parts if p and p.strip()]


def is_unidentified(raw):
    return normalize(raw) in UNIDENTIFIED_VALUES


# ---------------------------------------------------------------------------
# Carga y a­ndices de la base de Edimusica
# ---------------------------------------------------------------------------

def load_db(db_path):
    wb = openpyxl.load_workbook(db_path, read_only=True, data_only=True)
    ws = wb[DB_SHEET]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r is None:
            continue
        titulo, nomautor, codigo, codant, pct = (list(r) + [None] * 5)[:5]
        if not titulo:
            continue
        try:
            pctval = float(str(pct).replace(",", "."))
        except (TypeError, ValueError):
            pctval = 0.0
        rows.append({
            "titulo": titulo,
            "titulo_norm": normalize(titulo),
            "autor": nomautor,
            "autor_norm": normalize(nomautor),
            "codigo": codigo,
            "codant": codant,
            "pct": pctval,
        })
    wb.close()
    return rows


def build_exact_index(rows):
    idx = defaultdict(list)
    for i, r in enumerate(rows):
        idx[r["titulo_norm"]].append(i)
    return idx


def build_word_index(rows):
    doc_freq = defaultdict(int)
    word_rows = defaultdict(set)
    for i, r in enumerate(rows):
        words = {w for w in r["titulo_norm"].split()
                 if len(w) >= MIN_WORD_LEN_FOR_BLOCKING and w not in STOPWORDS}
        for w in words:
            doc_freq[w] += 1
            word_rows[w].add(i)
    return {w: idxs for w, idxs in word_rows.items() if doc_freq[w] <= MAX_DOC_FREQ_FOR_BLOCKING}


def get_fuzzy_candidates(title_norm, blocking_index, limit_pool=800):
    words = [w for w in title_norm.split()
             if len(w) >= MIN_WORD_LEN_FOR_BLOCKING and w not in STOPWORDS]
    candidate_ids = set()
    for w in words:
        if w in blocking_index:
            candidate_ids |= blocking_index[w]
            if len(candidate_ids) > limit_pool:
                break
    return candidate_ids


# ---------------------------------------------------------------------------
# Logica de matching
# ---------------------------------------------------------------------------

def summarize_group(rows, indices):
    total_pct = round(sum(rows[i]["pct"] for i in indices), 2)
    authors = [(rows[i]["autor"], rows[i]["pct"]) for i in indices]
    titulo = rows[indices[0]]["titulo"]
    codant = rows[indices[0]]["codant"]
    return {"total_pct": total_pct, "authors": authors, "titulo": titulo, "codant": codant}


def best_group_by_author(rows, candidate_indices, osa_authors_norm):
    by_group = defaultdict(list)
    for i in candidate_indices:
        by_group[rows[i]["codant"]].append(i)

    scored = []
    for codant, idxs in by_group.items():
        best_author_score = 0
        if osa_authors_norm:
            for i in idxs:
                a_norm = rows[i]["autor_norm"]
                for oa in osa_authors_norm:
                    sc = fuzz.token_sort_ratio(a_norm, oa)
                    if sc > best_author_score:
                        best_author_score = sc
        scored.append((best_author_score, codant, idxs))
    scored.sort(key=lambda x: -x[0])
    return scored


def match_row(rows, exact_index, blocking_index, titulo_raw, autor_raw):
    title_norm = normalize(titulo_raw)
    if not title_norm:
        return None

    osa_authors_norm = [a for a in (normalize(x) for x in split_authors(autor_raw)) if a]
    unidentified = is_unidentified(autor_raw)

    exact_candidates = exact_index.get(title_norm, [])

    if exact_candidates:
        scored_groups = best_group_by_author(rows, exact_candidates, osa_authors_norm)
        top_score, _, top_idxs = scored_groups[0]
        second_score = scored_groups[1][0] if len(scored_groups) > 1 else -1

        # El titulo exacto NO basta por si solo: titulos genericos/tradicionales
        # se repiten entre autores distintos. Siempre se exige confirmar el
        # autor (aun cuando solo haya un candidato), salvo que el autor OSA
        # venga "no identificado", caso en el que se manda a revision manual
        # con el autor sugerido en vez de asumirlo.
        if (not unidentified and top_score >= AUTHOR_CONFIRM_MIN
                and (top_score - second_score) >= AUTHOR_MARGIN_MIN):
            tier = "AUTO"
        else:
            tier = "REVISAR"

        return {
            "tier": tier,
            "title_score": 100,
            "author_score": top_score,
            "summary": summarize_group(rows, top_idxs),
            "match_type": "EXACTO",
        }

    candidate_ids = get_fuzzy_candidates(title_norm, blocking_index)
    if not candidate_ids:
        return None

    choices = {i: rows[i]["titulo_norm"] for i in candidate_ids}
    results = process.extract(
        title_norm, choices, scorer=fuzz.token_sort_ratio,
        limit=8, score_cutoff=FUZZY_CANDIDATE_CUTOFF,
    )
    if not results:
        return None

    matched_ids = [key for _, _, key in results]
    best_title_score = results[0][1]

    scored_groups = best_group_by_author(rows, matched_ids, osa_authors_norm)
    top_score, _, top_idxs = scored_groups[0]
    second_score = scored_groups[1][0] if len(scored_groups) > 1 else -1

    auto_ok = (
        best_title_score >= TIER1_FUZZY_TITLE_MIN
        and not unidentified
        and top_score >= AUTHOR_CONFIRM_MIN
        and (top_score - second_score) >= AUTHOR_MARGIN_MIN
    )
    tier = "AUTO" if auto_ok else "REVISAR"

    return {
        "tier": tier,
        "title_score": round(best_title_score, 1),
        "author_score": round(top_score, 1),
        "summary": summarize_group(rows, top_idxs),
        "match_type": "APROXIMADO",
    }


# ---------------------------------------------------------------------------
# Formato de comentarios
# ---------------------------------------------------------------------------

def format_authors(authors):
    return ", ".join(f"{a} ({p:.2f}%)" for a, p in authors)


def build_comment_auto(unidentified, summary):
    if not unidentified:
        return None
    return f"AUTOR(ES) IDENTIFICADO(S) SEGUN BD EDIMUSICA: {format_authors(summary['authors'])}"


def build_comment_review(match):
    s = match["summary"]
    return (
        f"REVISAR POSIBLE COINCIDENCIA EDIMUSICA ({match['match_type']}, "
        f"score titulo={match['title_score']}, score autor={match['author_score']}): "
        f"Titulo BD='{s['titulo']}' | Autor(es) BD: {format_authors(s['authors'])} "
        f"| % total BD={s['total_pct']:.2f} | Codigo={s['codant']}"
    )


# ---------------------------------------------------------------------------
# Programa principal
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Cruce CONSULTA OSA vs catalogo Edimusica")
    parser.add_argument("--osa", default=str(OSA_FILE_DEFAULT), help="Ruta al archivo de consulta de la OSA")
    parser.add_argument("--db", default=str(DB_FILE_DEFAULT), help="Ruta al archivo de obras de Edimusica")
    parser.add_argument("--limit", type=int, default=None, help="Procesar solo las primeras N filas (pruebas)")
    args = parser.parse_args()

    osa_path = Path(args.osa)
    db_path = Path(args.db)
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    output_path = osa_path.with_name(f"{osa_path.stem}_EDIMUSICA_{timestamp}.xlsx")
    log_path = osa_path.with_name(f"log_matching_{timestamp}.xlsx")

    t0 = time.time()
    print(f"Cargando base de datos Edimusica: {db_path}")
    rows = load_db(db_path)
    print(f"  {len(rows)} registros cargados.")

    print("Construyendo indices de busqueda...")
    exact_index = build_exact_index(rows)
    blocking_index = build_word_index(rows)
    print(f"  indice exacto: {len(exact_index)} titulos unicos")
    print(f"  indice de bloqueo: {len(blocking_index)} palabras")

    print(f"Abriendo consulta de la OSA: {osa_path}")
    wb = openpyxl.load_workbook(osa_path, data_only=False)
    ws = wb[OSA_SHEET]
    max_row = ws.max_row
    last_data_row = max_row if args.limit is None else min(max_row, OSA_FIRST_DATA_ROW + args.limit - 1)
    print(f"  {last_data_row - OSA_FIRST_DATA_ROW + 1} filas de datos a procesar.")

    log_wb = openpyxl.Workbook()
    log_ws = log_wb.active
    log_ws.title = "log"
    log_ws.append([
        "Fila OSA", "Titulo OSA", "Autor OSA", "Decision", "Tipo match",
        "Score titulo", "Score autor", "Titulo BD", "Autor(es) BD",
        "% total BD", "Codigo BD",
    ])

    n_auto = n_revisar = n_sin_match = n_ya_lleno = 0
    t1 = time.time()

    for row_idx in range(OSA_FIRST_DATA_ROW, last_data_row + 1):
        titulo_raw = ws.cell(row=row_idx, column=COL_TITULO).value
        autor_raw = ws.cell(row=row_idx, column=COL_AUTOR).value

        editor_cell = ws.cell(row=row_idx, column=COL_EDITOR)
        pct_cell = ws.cell(row=row_idx, column=COL_PCT)
        comentarios_cell = ws.cell(row=row_idx, column=COL_COMENTARIOS)

        if editor_cell.value not in (None, "") or pct_cell.value not in (None, ""):
            n_ya_lleno += 1
            continue

        if not titulo_raw:
            continue

        match = match_row(rows, exact_index, blocking_index, titulo_raw, autor_raw)

        if match is None:
            n_sin_match += 1
            continue

        unidentified = is_unidentified(autor_raw)
        summary = match["summary"]

        if match["tier"] == "AUTO":
            n_auto += 1
            editor_cell.value = EDITOR_NAME
            pct_cell.value = summary["total_pct"]
            if comentarios_cell.value in (None, ""):
                comment = build_comment_auto(unidentified, summary)
                if comment:
                    comentarios_cell.value = comment
            decision = "AUTO"
        else:
            n_revisar += 1
            if comentarios_cell.value in (None, ""):
                comentarios_cell.value = build_comment_review(match)
            decision = "REVISAR"

        log_ws.append([
            row_idx, titulo_raw, autor_raw, decision, match["match_type"],
            match["title_score"], match["author_score"], summary["titulo"],
            format_authors(summary["authors"]), summary["total_pct"], summary["codant"],
        ])

    t2 = time.time()
    print(f"Matching completado en {t2 - t1:.1f}s")
    print(f"  AUTO (llenado): {n_auto}")
    print(f"  REVISAR (comentario sugerido): {n_revisar}")
    print(f"  Sin coincidencia: {n_sin_match}")
    print(f"  Ya tenian EDITOR/% llenos (no tocadas): {n_ya_lleno}")

    wb.save(output_path)
    log_wb.save(log_path)
    print(f"Archivo de resultado: {output_path}")
    print(f"Log de auditoria:     {log_path}")
    print(f"Tiempo total: {time.time() - t0:.1f}s")


if __name__ == "__main__":
    main()
